"""
Excel Executive Report Exporter — 5-sheet workbook using openpyxl.

Sheet 1: Executive Summary (aggregated KPIs)
Sheet 2: Vendor Concentration (per-vendor exposure detail)
Sheet 3: Decision Log (full decision audit trail)
Sheet 4: Board Executive Summary (AI narrative + KPI dashboard)
Sheet 5: Price Shock Scenarios (color-coded EBITDA impact matrix)
"""

from io import BytesIO
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.exposure import FinancialExposure
from app.models.decision import Decision
from app.services.exposure_engine import calculate_all_exposures, EBITDA_MARGIN
from app.services.decision_store import DecisionStore


# --- Styling Constants ---
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _style_header_row(ws, num_cols: int):
    """Apply consistent header styling and freeze top row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _auto_size_columns(ws):
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _build_executive_summary(wb: Workbook, exposures: List[FinancialExposure], decisions: List[Decision]):
    """Sheet 1: Executive Summary KPIs."""
    ws = wb.active
    ws.title = "Executive Summary"

    headers = ["Metric", "Value"]
    ws.append(headers)

    total_exposure = sum(e.worst_case_exposure for e in exposures)
    total_spend = sum(e.annual_spend for e in exposures)
    high_risk_vendors = len([d for d in decisions if d.risk_level.value == "HIGH"])

    # Projected EBITDA impact = sum of all 10% shock ebitda deltas as baseline indicator
    projected_ebitda_impact = sum(e.estimated_ebitda_delta_10pct for e in exposures)

    total_opportunity = sum(d.annual_impact for d in decisions)

    kpis = [
        ("Total Identified Opportunity", total_opportunity),
        ("Total Vendor Spend", total_spend),
        ("Total Worst-Case Exposure", total_exposure),
        ("High Risk Vendors", high_risk_vendors),
        ("Projected EBITDA Impact (10% Shock)", projected_ebitda_impact),
        ("Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]

    for label, value in kpis:
        ws.append([label, value])

    # Format currency cells
    for row_idx in range(2, len(kpis) + 2):
        cell = ws.cell(row=row_idx, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = CURRENCY_FORMAT

    _style_header_row(ws, len(headers))
    _auto_size_columns(ws)


def _build_vendor_concentration(wb: Workbook, exposures: List[FinancialExposure]):
    """Sheet 2: Per-vendor concentration and exposure detail."""
    ws = wb.create_sheet("Vendor Concentration")

    headers = [
        "Vendor", "Category", "Annual Spend", "Share %",
        "Worst Case Exposure", "10% Shock Impact", "20% Shock Impact",
    ]
    ws.append(headers)

    for exp in sorted(exposures, key=lambda e: e.annual_spend, reverse=True):
        ws.append([
            exp.vendor_id,
            exp.category,
            exp.annual_spend,
            exp.vendor_share_pct,
            exp.worst_case_exposure,
            exp.price_shock_impact_10pct,
            exp.price_shock_impact_20pct,
        ])

    # Format columns
    for row_idx in range(2, ws.max_row + 1):
        # Currency columns: C, E, F, G (3, 5, 6, 7)
        for col_idx in [3, 5, 6, 7]:
            ws.cell(row=row_idx, column=col_idx).number_format = CURRENCY_FORMAT
        # Percentage column: D (4)
        ws.cell(row=row_idx, column=4).number_format = PERCENT_FORMAT

    _style_header_row(ws, len(headers))
    _auto_size_columns(ws)


def _build_decision_log(wb: Workbook, decisions: List[Decision]):
    """Sheet 3: Full decision log with events."""
    ws = wb.create_sheet("Decision Log")

    headers = [
        "Decision ID", "Vendor", "Recommendation", "Annual Impact",
        "Risk Level", "Status", "Created At",
    ]
    ws.append(headers)

    for d in decisions:
        ws.append([
            d.id,
            d.entity,
            d.recommended_action,
            d.annual_impact,
            d.risk_level.value,
            d.status.value,
            d.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Currency column: D (4)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = CURRENCY_FORMAT

    _style_header_row(ws, len(headers))
    _auto_size_columns(ws)


# ── Board Report Styling Constants ───────────────────────────────────
NAVY = "0F172A"
BLUE = "2563EB"
LIGHT_BLUE = "DBEAFE"
GREEN = "16A34A"
AMBER = "B45309"
RED = "DC2626"
LIGHT_GREEN = "DCFCE7"
LIGHT_AMBER = "FEF3C7"
LIGHT_RED = "FEE2E2"
LIGHT_GREY = "F8FAFC"
WHITE = "FFFFFF"


def _build_board_executive_summary(
    wb: Workbook,
    narrative: Optional[str],
    total_spend: float,
    high_risk_count: int,
    medium_risk_count: int,
    decision_count: int,
    estimated_savings: float,
    ebitda_at_risk: float,
    vendor_count: int,
):
    """Sheet 4: Board Executive Summary with AI narrative and KPI dashboard."""
    ws = wb.create_sheet("Board Executive Summary")

    today_str = datetime.utcnow().strftime("%B %d, %Y")

    # ── Title Block ──
    ws["A1"] = "CADE — CAPITAL ALLOCATION DECISION ENGINE"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)

    ws["A2"] = f"Board Report — {today_str}"
    ws["A2"].font = Font(name="Calibri", size=11, color="64748B")

    # Row 3: blank

    # ── Section Header ──
    ws["A4"] = "Executive Summary"
    ws["A4"].font = Font(name="Calibri", bold=True, size=12, color=NAVY)

    # ── AI Narrative Block (A5:F10) ──
    narrative_text = narrative or "AI narrative unavailable — no data or API key configured."
    ws.merge_cells("A5:F10")
    cell = ws["A5"]
    cell.value = narrative_text
    cell.font = Font(name="Calibri", size=11)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    blue_border = Border(
        left=Side(style="thin", color=BLUE),
        right=Side(style="thin", color=BLUE),
        top=Side(style="thin", color=BLUE),
        bottom=Side(style="thin", color=BLUE),
    )
    cell.border = blue_border

    # Row 11: blank

    # ── KPI Table Headers (Row 12) ──
    kpi_header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    kpi_header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    kpi_header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header_text in enumerate(["KPI", "Value", "Notes"], start=1):
        c = ws.cell(row=12, column=col_idx, value=header_text)
        c.font = kpi_header_font
        c.fill = kpi_header_fill
        c.alignment = kpi_header_align

    # ── KPI Data Rows (13–19), alternating fills ──
    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    grey_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")

    kpi_rows = [
        ("Total Vendor Spend", f"${total_spend:,.0f}", "Across all categories"),
        ("HIGH Risk Vendors", str(high_risk_count), "Require immediate review"),
        ("MEDIUM Risk Vendors", str(medium_risk_count), "Monitor closely"),
        ("Decisions Generated", str(decision_count), "This period"),
        ("Estimated Savings", f"${estimated_savings:,.0f}", "If actions implemented"),
        ("EBITDA at Risk (10%)", f"${ebitda_at_risk:,.0f}", "10% price shock scenario"),
        ("Vendors Tracked", str(vendor_count), "Active this period"),
    ]

    for i, (kpi, value, note) in enumerate(kpi_rows):
        row_idx = 13 + i
        fill = white_fill if i % 2 == 0 else grey_fill
        for col_idx, text in enumerate([kpi, value, note], start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=text)
            c.fill = fill
            c.font = Font(name="Calibri", size=11)

    # ── Column Widths ──
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35


def _get_impact_fill_and_font(impact: float):
    """Return (PatternFill, Font) based on impact magnitude."""
    if impact < 10_000:
        return (
            PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"),
            Font(name="Calibri", size=11, color=GREEN),
        )
    elif impact < 50_000:
        return (
            PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid"),
            Font(name="Calibri", size=11, color=AMBER),
        )
    else:
        return (
            PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"),
            Font(name="Calibri", size=11, color=RED),
        )


def _build_price_shock_scenarios(
    wb: Workbook,
    exposures: List[FinancialExposure],
    decisions: List[Decision],
):
    """Sheet 5: Price Shock Impact Analysis with color-coded EBITDA impacts."""
    ws = wb.create_sheet("Price Shock Scenarios")

    # ── Title ──
    ws["A1"] = "Price Shock Impact Analysis"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=NAVY)

    ws["A2"] = "EBITDA impact if vendors raise prices"
    ws["A2"].font = Font(name="Calibri", size=11, color="64748B")

    # Row 3: blank

    # ── Headers (Row 4) ──
    headers = [
        "Vendor", "Category", "Annual Spend",
        "5% Impact", "10% Impact", "15% Impact", "20% Impact",
        "Risk Level",
    ]
    shock_header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    shock_header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = shock_header_font
        c.fill = shock_header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Build vendor → risk_level lookup from decisions
    vendor_risk = {}
    for d in decisions:
        existing = vendor_risk.get(d.entity)
        if existing is None or d.risk_level.value == "HIGH":
            vendor_risk[d.entity] = d.risk_level.value

    margin = EBITDA_MARGIN
    shock_pcts = [0.05, 0.10, 0.15, 0.20]

    total_spend = 0.0
    total_impacts = [0.0, 0.0, 0.0, 0.0]

    sorted_exposures = sorted(exposures, key=lambda e: e.annual_spend, reverse=True)

    for row_offset, exp in enumerate(sorted_exposures):
        row_idx = 5 + row_offset
        spend = exp.annual_spend
        total_spend += spend

        # Vendor + Category + Spend
        ws.cell(row=row_idx, column=1, value=exp.vendor_id)
        ws.cell(row=row_idx, column=2, value=exp.category)
        spend_cell = ws.cell(row=row_idx, column=3, value=spend)
        spend_cell.number_format = CURRENCY_FORMAT

        # Shock impacts (columns D-G)
        for i, pct in enumerate(shock_pcts):
            impact = spend * pct * margin
            total_impacts[i] += impact
            c = ws.cell(row=row_idx, column=4 + i, value=impact)
            c.number_format = CURRENCY_FORMAT
            fill, font = _get_impact_fill_and_font(impact)
            c.fill = fill
            c.font = font

        # Risk Level (column H)
        risk = vendor_risk.get(exp.vendor_id, "—")
        ws.cell(row=row_idx, column=8, value=risk)

    # ── TOTAL Row ──
    total_row = 5 + len(sorted_exposures)
    bold_font = Font(name="Calibri", bold=True, size=11)

    ws.cell(row=total_row, column=1, value="TOTAL PORTFOLIO").font = bold_font
    total_spend_cell = ws.cell(row=total_row, column=3, value=total_spend)
    total_spend_cell.number_format = CURRENCY_FORMAT
    total_spend_cell.font = bold_font

    for i, total in enumerate(total_impacts):
        c = ws.cell(row=total_row, column=4 + i, value=total)
        c.number_format = CURRENCY_FORMAT
        c.font = bold_font

    # ── Column Widths ──
    col_widths = {"A": 25, "B": 20, "C": 18, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w


def generate_executive_report(
    decisions: Optional[List[Decision]] = None,
    exposures: Optional[List[FinancialExposure]] = None,
    narrative: Optional[str] = None,
    total_spend: float = 0.0,
    high_risk_count: int = 0,
    medium_risk_count: int = 0,
    decision_count: int = 0,
    estimated_savings: float = 0.0,
    ebitda_at_risk: float = 0.0,
    vendor_count: int = 0,
) -> BytesIO:
    """
    Generate the full 5-sheet Executive Report workbook.
    Returns a BytesIO buffer containing the .xlsx file.

    If decisions/exposures are not provided, they are loaded internally
    (backward-compatible with original callers).
    """
    if exposures is None:
        exposures = calculate_all_exposures()
    if decisions is None:
        decisions = DecisionStore.get_all_decisions()

    # Enrich decisions with events
    for d in decisions:
        d.events = DecisionStore.get_events_for_decision(d.id)

    wb = Workbook()

    # Original 3 sheets (untouched)
    _build_executive_summary(wb, exposures, decisions)
    _build_vendor_concentration(wb, exposures)
    _build_decision_log(wb, decisions)

    # New Board Report sheets
    _build_board_executive_summary(
        wb,
        narrative=narrative,
        total_spend=total_spend,
        high_risk_count=high_risk_count,
        medium_risk_count=medium_risk_count,
        decision_count=decision_count,
        estimated_savings=estimated_savings,
        ebitda_at_risk=ebitda_at_risk,
        vendor_count=vendor_count,
    )
    _build_price_shock_scenarios(wb, exposures, decisions)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

"""
Procurement Intelligence API — Supplier price comparison, bulk buy alerts.

GET /api/procurement/price-comparison         → all categories
GET /api/procurement/price-comparison/{cat}    → single category
GET /api/procurement/bulk-buy-alerts           → bulk buy opportunities
"""
from datetime import date

from fastapi import APIRouter, HTTPException
from app.services.price_comparison_engine import PriceComparisonEngine
from app.services.item_price_engine import ItemPriceEngine

router = APIRouter()


@router.get("/price-comparison")
async def get_price_comparison():
    """Compare supplier pricing across all categories with 2+ suppliers."""
    results = PriceComparisonEngine.analyze_all_categories()

    total_overspend = sum(r.annual_overspend for r in results)
    total_savings = sum(r.estimated_annual_savings for r in results)
    categories_with_multiple = len([
        r for r in results if len(r.suppliers) >= 2
    ])

    return {
        "categories_analyzed": len(results),
        "categories_with_multiple_suppliers": categories_with_multiple,
        "total_annual_overspend": round(total_overspend, 2),
        "total_estimated_savings": round(total_savings, 2),
        "comparisons": [
            {
                "category": r.item_category,
                "supplier_count": len(r.suppliers),
                "suppliers": [
                    {
                        "name": s.supplier_name,
                        "avg_purchase_amount": round(s.unit_price, 2),
                        "total_purchased": round(s.total_purchased, 2),
                        "purchase_count": s.purchase_count,
                        "price_trend": s.price_trend,
                        "reliability_score": round(
                            s.reliability_score, 2
                        ),
                    }
                    for s in sorted(
                        r.suppliers,
                        key=lambda x: x.unit_price
                    )
                ],
                "cheapest_supplier": r.cheapest_supplier,
                "price_variance_pct": r.price_variance_pct,
                "annual_overspend": r.annual_overspend,
                "recommended_primary": r.recommended_primary_supplier,
                "recommended_backup": r.recommended_backup_supplier,
                "estimated_savings": r.estimated_annual_savings,
                "bulk_buy_recommended": r.bulk_buy_recommended,
                "bulk_buy_reasoning": r.bulk_buy_reasoning,
                "consolidation_recommended": r.consolidation_recommended,
            }
            for r in results
        ]
    }


@router.get("/price-comparison/{category}")
async def get_category_price_comparison(category: str):
    """Compare suppliers within a single procurement category."""
    result = PriceComparisonEngine.analyze_category_suppliers(
        category.replace("-", " ")
    )
    if not result:
        raise HTTPException(
            404,
            f"No multi-supplier data found for category: {category}"
        )
    return {
        "category": result.item_category,
        "supplier_count": len(result.suppliers),
        "suppliers": [
            {
                "name": s.supplier_name,
                "avg_purchase_amount": round(s.unit_price, 2),
                "total_purchased": round(s.total_purchased, 2),
                "purchase_count": s.purchase_count,
                "price_trend": s.price_trend,
                "reliability_score": round(s.reliability_score, 2),
            }
            for s in sorted(result.suppliers, key=lambda x: x.unit_price)
        ],
        "cheapest_supplier": result.cheapest_supplier,
        "price_variance_pct": result.price_variance_pct,
        "annual_overspend": result.annual_overspend,
        "recommended_primary": result.recommended_primary_supplier,
        "recommended_backup": result.recommended_backup_supplier,
        "estimated_savings": result.estimated_annual_savings,
        "bulk_buy_recommended": result.bulk_buy_recommended,
        "bulk_buy_reasoning": result.bulk_buy_reasoning,
        "consolidation_recommended": result.consolidation_recommended,
    }


@router.get("/bulk-buy-alerts")
async def get_bulk_buy_alerts():
    """Identify categories where bulk purchasing is recommended due to rising prices."""
    results = PriceComparisonEngine.analyze_all_categories()
    alerts = [r for r in results if r.bulk_buy_recommended]

    return {
        "total_alerts": len(alerts),
        "total_savings_opportunity": round(
            sum(r.estimated_annual_savings for r in alerts), 2
        ),
        "alerts": [
            {
                "category": r.item_category,
                "reasoning": r.bulk_buy_reasoning,
                "estimated_savings": r.estimated_annual_savings,
                "recommended_supplier": r.recommended_primary_supplier,
            }
            for r in alerts
        ]
    }


@router.get("/item-price-mismatches")
async def get_item_price_mismatches():
    """
    Find items where multiple suppliers sell the SAME item at different prices.
    Returns switching recommendations with per-item savings calculations.
    """
    return ItemPriceEngine.get_summary()


@router.get("/bulk-buy-recommendations")
async def get_bulk_buy_recommendations():
    """Analyze purchase patterns and recommend bulk buying opportunities."""
    from app.services.bulk_buy_engine import BulkBuyEngine
    from dataclasses import asdict

    recs = BulkBuyEngine.analyze()
    total_saving = sum(r.net_saving for r in recs)
    urgent = [r for r in recs if r.urgency == "BUY NOW"]

    return {
        "total_recommendations": len(recs),
        "urgent_count": len(urgent),
        "total_net_saving": round(total_saving, 2),
        "recommendations": [asdict(r) for r in recs],
    }


@router.get("/price-comparison/report")
async def download_price_mismatch_report():
    """Generate Excel report of all price mismatches across categories."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    results = PriceComparisonEngine.analyze_all_categories()
    item_mismatches = ItemPriceEngine.get_summary()

    wb = Workbook()

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active
    ws1.title = "Executive Summary"

    header_font = Font(name="Calibri", size=14, bold=True, color="1a365d")
    sub_font = Font(name="Calibri", size=10, color="6b7280")
    bold_font = Font(name="Calibri", size=11, bold=True)
    navy_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    white_font = Font(name="Calibri", size=10, bold=True, color="ffffff")
    red_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    amber_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")

    ws1["A1"] = "CADE Hospital — Price Mismatch Report"
    ws1["A1"].font = header_font
    ws1["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws1["A2"].font = sub_font
    ws1.merge_cells("A1:F1")

    total_overspend = sum(r.annual_overspend for r in results)
    total_savings = sum(r.estimated_annual_savings for r in results)
    severe = [r for r in results if r.price_variance_pct >= 30]
    moderate = [r for r in results if 15 <= r.price_variance_pct < 30]

    summary_data = [
        ("Total categories analysed", len(results)),
        ("Categories with price mismatches", len([r for r in results if r.price_variance_pct > 0])),
        ("Severe mismatches (30%+)", len(severe)),
        ("Moderate mismatches (15-30%)", len(moderate)),
        ("Total annual savings opportunity", f"₹{total_savings:,.0f}"),
    ]

    row = 4
    for label, val in summary_data:
        ws1[f"A{row}"] = label
        ws1[f"A{row}"].font = bold_font
        ws1[f"C{row}"] = val
        row += 1

    # Top 5 savings
    row += 1
    ws1[f"A{row}"] = "Top 5 Savings Opportunities"
    ws1[f"A{row}"].font = header_font
    row += 1
    headers = ["Category", "Overspend", "Cheapest Supplier", "Savings"]
    for i, h in enumerate(headers):
        cell = ws1.cell(row=row, column=i+1, value=h)
        cell.font = white_font
        cell.fill = navy_fill
    row += 1
    for r in sorted(results, key=lambda x: x.estimated_annual_savings, reverse=True)[:5]:
        ws1.cell(row=row, column=1, value=r.item_category)
        ws1.cell(row=row, column=2, value=round(r.annual_overspend, 0))
        ws1.cell(row=row, column=3, value=r.cheapest_supplier)
        ws1.cell(row=row, column=4, value=round(r.estimated_annual_savings, 0))
        row += 1

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 25
    ws1.column_dimensions["D"].width = 15

    # ── Sheet 2: All Mismatches ──
    ws2 = wb.create_sheet("All Mismatches")
    headers2 = [
        "Category", "Supplier Count", "Cheapest Supplier",
        "Price Variance %", "Annual Overspend", "Estimated Savings",
        "Recommended Primary", "Bulk Buy?",
    ]
    for i, h in enumerate(headers2):
        cell = ws2.cell(row=1, column=i+1, value=h)
        cell.font = white_font
        cell.fill = navy_fill

    for idx, r in enumerate(sorted(results, key=lambda x: x.annual_overspend, reverse=True), start=2):
        ws2.cell(row=idx, column=1, value=r.item_category)
        ws2.cell(row=idx, column=2, value=len(r.suppliers))
        ws2.cell(row=idx, column=3, value=r.cheapest_supplier)
        ws2.cell(row=idx, column=4, value=r.price_variance_pct)
        ws2.cell(row=idx, column=5, value=round(r.annual_overspend, 0))
        ws2.cell(row=idx, column=6, value=round(r.estimated_annual_savings, 0))
        ws2.cell(row=idx, column=7, value=r.recommended_primary_supplier)
        ws2.cell(row=idx, column=8, value="Yes" if r.bulk_buy_recommended else "No")

        if r.annual_overspend > 500000:
            for c in range(1, 9):
                ws2.cell(row=idx, column=c).fill = red_fill
        elif r.annual_overspend > 100000:
            for c in range(1, 9):
                ws2.cell(row=idx, column=c).fill = amber_fill

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws2.column_dimensions[col].width = 20

    # ── Sheet 3: By Category ──
    ws3 = wb.create_sheet("By Category")
    cat_headers = ["Category", "Vendor Count", "Total Overspend", "Recommended Supplier"]
    for i, h in enumerate(cat_headers):
        cell = ws3.cell(row=1, column=i+1, value=h)
        cell.font = white_font
        cell.fill = navy_fill
    for idx, r in enumerate(results, start=2):
        ws3.cell(row=idx, column=1, value=r.item_category)
        ws3.cell(row=idx, column=2, value=len(r.suppliers))
        ws3.cell(row=idx, column=3, value=round(r.annual_overspend, 0))
        ws3.cell(row=idx, column=4, value=r.recommended_primary_supplier)
    for col in ["A", "B", "C", "D"]:
        ws3.column_dimensions[col].width = 25

    # ── Sheet 4: Supplier Rankings ──
    ws4 = wb.create_sheet("Supplier Rankings")
    rank_headers = ["Category", "Rank", "Supplier", "Avg Price", "Purchases", "Total Spend", "Reliability", "Recommendation"]
    for i, h in enumerate(rank_headers):
        cell = ws4.cell(row=1, column=i+1, value=h)
        cell.font = white_font
        cell.fill = navy_fill

    row = 2
    for r in results:
        for rank, s in enumerate(sorted(r.suppliers, key=lambda x: x.unit_price), start=1):
            rec = "Primary" if s.supplier_name == r.recommended_primary_supplier else (
                "Backup" if s.supplier_name == r.recommended_backup_supplier else ""
            )
            ws4.cell(row=row, column=1, value=r.item_category)
            ws4.cell(row=row, column=2, value=rank)
            ws4.cell(row=row, column=3, value=s.supplier_name)
            ws4.cell(row=row, column=4, value=round(s.unit_price, 0))
            ws4.cell(row=row, column=5, value=s.purchase_count)
            ws4.cell(row=row, column=6, value=round(s.total_purchased, 0))
            ws4.cell(row=row, column=7, value=f"{s.reliability_score:.0%}")
            ws4.cell(row=row, column=8, value=rec)
            row += 1
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws4.column_dimensions[col].width = 18

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"price_mismatch_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/score")
async def get_procurement_score():
    """
    Calculate overall hospital procurement health score (0-100).
    5 components: Price Competitiveness, Vendor Diversification,
    Contract Management, Spend Control, Savings Capture.
    """
    from app.services.exposure_engine import calculate_all_exposures
    from app.services.decision_store import DecisionStore
    from app.models.decision import DecisionStatus

    exposures = calculate_all_exposures()
    results = PriceComparisonEngine.analyze_all_categories()
    decisions = DecisionStore.get_all_decisions()

    # ── 1. Price Competitiveness (0-100) ──
    # % of categories where we're using the cheapest supplier
    if results:
        competitive_count = sum(
            1 for r in results if r.price_variance_pct < 15
        )
        price_score = (competitive_count / len(results)) * 100
    else:
        price_score = 50

    # ── 2. Vendor Diversification (0-100) ──
    # Based on average category concentration
    if exposures:
        avg_concentration = sum(
            e.vendor_share_pct for e in exposures
        ) / len(exposures)
        # Lower concentration = better score
        diversification_score = max(0, min(100, (1 - avg_concentration) * 100))
    else:
        diversification_score = 50

    # ── 3. Contract Management (0-100) ──
    # Based on how many renewals are handled proactively
    try:
        from app.api.contracts import get_renewal_date
        import hashlib
        total_vendors = len(exposures)
        proactive = sum(
            1 for e in exposures
            if (get_renewal_date(e.vendor_id) - date.today()).days > 30
        )
        from datetime import date
        contract_score = (proactive / max(total_vendors, 1)) * 100
    except Exception:
        contract_score = 60

    # ── 4. Spend Control (0-100) ──
    # Vendors within threshold vs total
    try:
        from app.services.policy_engine import policy_engine
        within_threshold = 0
        for e in exposures:
            pol = policy_engine.get_policy(e.category)
            threshold = pol.get("spend_threshold", 100000)
            if e.annual_spend <= threshold:
                within_threshold += 1
        spend_score = (within_threshold / max(len(exposures), 1)) * 100
    except Exception:
        spend_score = 50

    # ── 5. Savings Capture (0-100) ──
    total_decisions = len(decisions)
    approved = len([d for d in decisions if d.status == DecisionStatus.APPROVED])
    savings_score = (approved / max(total_decisions, 1)) * 100 if total_decisions > 0 else 0

    # ── Overall Score ──
    overall = (
        price_score * 0.25 +
        diversification_score * 0.20 +
        contract_score * 0.20 +
        spend_score * 0.20 +
        savings_score * 0.15
    )

    # Grade
    if overall >= 80:
        grade = "A"
    elif overall >= 60:
        grade = "B"
    elif overall >= 40:
        grade = "C"
    elif overall >= 20:
        grade = "D"
    else:
        grade = "F"

    # Top 3 improvement actions  
    improvements = []
    component_scores = [
        ("Price Competitiveness", price_score, "Negotiate better rates with top spend vendors", 8),
        ("Vendor Diversification", diversification_score, "Qualify alternative suppliers for concentrated categories", 10),
        ("Contract Management", contract_score, "Set up proactive renewal tracking 90 days ahead", 12),
        ("Spend Control", spend_score, "Implement purchase order approval for above-threshold vendors", 7),
        ("Savings Capture", savings_score, "Review and act on pending procurement decisions", 15),
    ]
    component_scores.sort(key=lambda x: x[1])
    for name, score, action, improvement in component_scores[:3]:
        improvements.append({
            "component": name,
            "current_score": round(score, 1),
            "action": action,
            "expected_improvement": improvement,
            "effort": "Medium" if improvement < 10 else "High",
        })

    return {
        "overall_score": round(overall, 1),
        "grade": grade,
        "components": {
            "price_competitiveness": round(price_score, 1),
            "vendor_diversification": round(diversification_score, 1),
            "contract_management": round(contract_score, 1),
            "spend_control": round(spend_score, 1),
            "savings_capture": round(savings_score, 1),
        },
        "benchmarks": {
            "your_score": round(overall, 1),
            "industry_average": 58,
            "top_quartile": 75,
        },
        "improvement_roadmap": improvements,
    }

